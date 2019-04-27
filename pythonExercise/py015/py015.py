# coding=utf-8

# https://github.com/Show-Me-the-Code/show-me-the-code

"""
纯文本文件 city.txt为城市信息, 里面的内容（包括花括号）如下所示：

{
    "1" : "上海",
    "2" : "北京",
    "3" : "成都"
}
请将上述内容写到 city.xls 文件中，如下图所示：

city.xls
"""

import openpyxl
import json


def open_json_file(file_path):
    with open(file_path, encoding="utf-8", mode="r") as f:
        json_data = f.read()
        return json.loads(json_data)


# print(open_json_file(file_path))


def txt2xlsx(xlsx_path, sheet_name):
    '''

    :param xlsx_path: path of xlsx file
    :param sheet_name: sheet name
    :return: workbook and worksheet obj
    '''

    workbook = openpyxl.load_workbook(xlsx_path)
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)

    data = open_json_file(file_path)

    rows_total = len(data)
    for row in range(1, rows_total+1):
        worksheet.cell(row=row, column=1).value = row
        worksheet.cell(row=row, column=2).value = data[str(row)]

    workbook.save(xlsx_path)


if __name__ == '__main__':
    file_path = "./city.txt"
    xlsx_path = "./test.xlsx"
    txt2xlsx(xlsx_path, "AAA")
